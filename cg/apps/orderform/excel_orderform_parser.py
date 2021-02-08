import logging
from pathlib import Path
from typing import Dict, List

import openpyxl
from cg.apps.orderform.orderform_parser import OrderformParser
from cg.apps.orderform.orderform_schema import OrderformSchema
from cg.exc import OrderFormError
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LOG = logging.getLogger(__name__)


class ExcelOrderformParser(OrderformParser):
    SHEET_NAMES: List[str] = ["Orderform", "orderform", "order form"]
    VALID_ORDERFORMS: List[str] = [
        "1508:22",  # Orderform MIP, Balsamic, sequencing only, MIP RNA
        "1541:6",  # Orderform Externally sequenced samples
        "1603:9",  # Microbial WGS
        "1604:10",  # Orderform Ready made libraries (RML)
        "1605:8",  # Microbial metagenomes
    ]

    def check_orderform_version(self, document_title: str) -> None:
        """Raise an error if the orderform is too new or too old for the order portal."""
        for valid_orderform in self.VALID_ORDERFORMS:
            if valid_orderform in document_title:
                return
        raise OrderFormError(f"Unsupported orderform: {document_title}")

    def get_sheet_name(self, sheet_names: List[str]) -> str:
        """Return the correct (exitsing) sheet names"""

        for name in self.SHEET_NAMES:
            if name not in sheet_names:
                continue
            LOG.info("Found sheet name %s", name)
            return name
        raise OrderFormError("'orderform' sheet not found in Excel file")

    def get_document_title(self, workbook: Workbook, orderform_sheet: Worksheet) -> str:
        """Get the document title for the order form.

        Openpyxl use 1 based counting
        """
        for sheet_number, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name.lower() != "information":
                continue
            information_sheet: Worksheet = workbook[sheet_name]
            document_title = information_sheet.cell(1, 3).value
            LOG.info("Found document title %s", document_title)
            return document_title

        document_title = orderform_sheet.cell(1, 2).value
        LOG.info("Found document title %s", document_title)
        return document_title

    def relevant_rows(self, orderform_sheet: Worksheet) -> List[Dict[str, str]]:
        """Get the relevant rows from an order form sheet."""
        raw_samples = []
        current_row = None
        empty_row_found = False
        row: tuple
        for row in orderform_sheet.rows:
            if row[0].value == "</SAMPLE ENTRIES>":
                LOG.debug("End of samples info")
                break

            if current_row == "header":
                header_row = [cell.value for cell in row]
                current_row = None
            elif current_row == "samples":
                values = []
                for cell in row:
                    value = str(cell.value)
                    if value == "None":
                        value = ""
                    values.append(value)

                # skip empty rows
                if values[0]:
                    if empty_row_found:
                        raise OrderFormError(
                            f"Found data after empty lines. Please delete any "
                            f"non-sample data rows in between the samples"
                        )

                    sample_dict = dict(zip(header_row, values))
                    raw_samples.append(sample_dict)
                else:
                    empty_row_found = True

            if row[0].value == "<TABLE HEADER>":
                LOG.debug("Found header row")
                current_row = "header"
            elif row[0].value == "<SAMPLE ENTRIES>":
                LOG.debug("Found samples row")
                current_row = "samples"
        return raw_samples

    def parse_sample(self, raw_sample: Dict[str, str]) -> dict:
        """Parse a raw sample row from order form sheet."""
        if ":" in raw_sample.get("UDF/Gene List", ""):
            raw_sample["UDF/Gene List"] = raw_sample["UDF/Gene List"].replace(":", ";")

        if raw_sample["UDF/priority"].lower() == "förtur":
            raw_sample["UDF/priority"] = "priority"
        raw_source = raw_sample.get("UDF/Source")
        sample = {
            "application": raw_sample["UDF/Sequencing Analysis"],
            "capture_kit": raw_sample.get("UDF/Capture Library version"),
            "case": raw_sample.get("UDF/familyID"),
            "comment": raw_sample.get("UDF/Comment"),
            "container": raw_sample.get("Container/Type"),
            "container_name": raw_sample.get("Container/Name"),
            "custom_index": raw_sample.get("UDF/Custom index"),
            "customer": raw_sample["UDF/customer"],
            "data_analysis": raw_sample["UDF/Data Analysis"],
            "data_delivery": raw_sample.get("UDF/Data Delivery"),
            "elution_buffer": raw_sample.get("UDF/Sample Buffer"),
            "extraction_method": raw_sample.get("UDF/Extraction method"),
            "formalin_fixation_time": raw_sample.get("UDF/Formalin Fixation Time"),
            "index": raw_sample.get("UDF/Index type"),
            "from_sample": raw_sample.get("UDF/is_for_sample"),
            "name": raw_sample["Sample/Name"],
            "organism": raw_sample.get("UDF/Strain"),
            "organism_other": raw_sample.get("UDF/Other species"),
            "panels": (
                raw_sample["UDF/Gene List"].split(";") if raw_sample.get("UDF/Gene List") else None
            ),
            "pool": raw_sample.get("UDF/pool name"),
            "post_formalin_fixation_time": raw_sample.get("UDF/Post Formalin Fixation Time"),
            "priority": raw_sample["UDF/priority"].lower()
            if raw_sample.get("UDF/priority")
            else None,
            "reagent_label": raw_sample.get("Sample/Reagent Label"),
            "reference_genome": raw_sample.get("UDF/Reference Genome Microbial"),
            "require_qcok": raw_sample.get("UDF/Process only if QC OK") == "yes",
            "rml_plate_name": raw_sample.get("UDF/RML plate name"),
            "sex": REV_SEX_MAP.get(raw_sample.get("UDF/Gender", "").strip()),
            "source": raw_source if raw_source in SOURCE_TYPES else None,
            "status": raw_sample["UDF/Status"].lower() if raw_sample.get("UDF/Status") else None,
            "tissue_block_size": raw_sample.get("UDF/Tissue Block Size"),
            "tumour": raw_sample.get("UDF/tumor") == "yes",
            "tumour_purity": raw_sample.get("UDF/tumour purity"),
            "well_position": raw_sample.get("Sample/Well Location"),
            "well_position_rml": raw_sample.get("UDF/RML well position"),
        }

        numeric_attributes = [
            ("index_number", "UDF/Index number"),
            ("volume", "UDF/Volume (uL)"),
            ("quantity", "UDF/Quantity"),
            ("concentration", "UDF/Concentration (nM)"),
            ("concentration_sample", "UDF/Sample Conc."),
            ("time_point", "UDF/time_point"),
        ]
        for json_key, excel_key in numeric_attributes:
            str_value = raw_sample.get(excel_key, "").rsplit(".0")[0]
            if str_value.replace(".", "").isnumeric():
                sample[json_key] = str_value

        for parent in ["mother", "father"]:
            parent_key = f"UDF/{parent}ID"
            sample[parent] = (
                raw_sample[parent_key]
                if raw_sample.get(parent_key) and (raw_sample[parent_key] != "0.0")
                else None
            )

        return sample

    def parse_orderform(self, excel_path: str) -> OrderformSchema:
        """Parse out information from an order form."""

        workbook: Workbook = openpyxl.load_workbook(
            filename=excel_path, read_only=True, data_only=True
        )

        sheet_name: str = self.get_sheet_name(workbook.sheetnames)

        orderform_sheet: Worksheet = workbook[sheet_name]
        document_title: str = self.get_document_title(
            workbook=workbook, orderform_sheet=orderform_sheet
        )
        self.check_orderform_version(document_title)

        raw_samples = self.relevant_rows(orderform_sheet)

        if len(raw_samples) == 0:
            raise OrderFormError("orderform doesn't contain any samples")
        parsed_samples = [parse_sample(raw_sample) for raw_sample in raw_samples]
        #
        # project_type = get_project_type(document_title, parsed_samples)
        # delivery_type = get_data_delivery(parsed_samples, OrderType(project_type))
        #
        # if project_type in CASE_PROJECT_TYPES:
        #     parsed_cases = group_cases(parsed_samples)
        #     items = []
        #     customer_ids = set()
        #     for case_id, parsed_case in parsed_cases.items():
        #         customer_id, case_data = expand_case(case_id, parsed_case, delivery_type)
        #         customer_ids.add(customer_id)
        #         items.append(case_data)
        # else:
        #     customer_ids = set(sample["customer"] for sample in parsed_samples)
        #     items = parsed_samples
        #
        # customer_options = len(customer_ids)
        # if customer_options == 0:
        #     raise OrderFormError("Customer information is missing")
        # elif customer_options != 1:
        #     raise OrderFormError(f"Samples have different customers: {customer_ids}")
        #
        # filename_base = Path(excel_path).stem
        # parsed_order = {
        #     "customer": customer_ids.pop(),
        #     "delivery_type": delivery_type,
        #     "items": items,
        #     "name": filename_base,
        #     "project_type": project_type,
        # }

        return parsed_order
